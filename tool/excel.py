import os
import time

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


class Excel:

    def __init__(self, filename):
        self.filename = filename
        self.base_dir = os.path.dirname(os.path.dirname(__file__))
        self.excel_path = os.sep.join([self.base_dir, 'data', filename])
        name_list = self.filename.split('.')
        self.save_name = name_list[0] + '_' + time.strftime('%m%d%H%M') + '.' + name_list[1]

    def get_workbook(self):
        return load_workbook(self.excel_path)

    def get_excel_dir(self):
        save_path = os.sep.join([self.base_dir, 'data', self.save_name])
        return save_path

    def get_excel_cases(self):
        wb = self.get_workbook()
        ws = wb['测试用例集合']
        case_list = []
        for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
            row_list = [cell.value for cell in row_cells]
            if row_list[0]:
                case_list.append(row_list[1])
        return case_list

    def get_excel_case_operation(self, case):
        """如果要测试结果回显，这个方法就不能用了"""
        ws = self.wb[case]
        case_op_list = []
        for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
            row_list = [cell.value for cell in row_cells]
            if row_list[0]:
                case_op_list.append({
                    'operation': row_list[0],
                    'element': row_list[1],
                    'data': row_list[2]
                })
        return case_op_list
    #
    # def get_excel_elements(self):
    #     wb = self.get_workbook()
    #     ws = wb['测试元素库']
    #     for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
    #         row_list = [cell.value for cell in row_cells]
    #         if row_list[0]:
    #             ELEMENT[row_list[0]] = (row_list[1], row_list[2])


if __name__ == '__main__':
    excel = Excel('mtx_shop.xlsx')
    wb = excel.get_workbook()
    ws = wb['登录']
    ws.cell(2, 4, 'PASS').fill = PatternFill("solid", fgColor='5FB878')
    wb.save('mtx_shop.xlsx')